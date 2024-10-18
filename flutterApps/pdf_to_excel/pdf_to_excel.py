import tkinter as tk
from tkinter import filedialog, ttk
import PyPDF2
import os
import threading
import time
import subprocess
from openpyxl import Workbook
import re

class PDFConverter:
    def __init__(self, master):
        self.master = master
        self.master.title("PDF to Excel Converter")
        self.master.geometry("300x250")

        self.pdf_file = None
        self.excel_file = None
        self.conversion_thread = None
        self.cancel_flag = threading.Event()

        self.import_button = tk.Button(master, text="Import", command=self.import_pdf)
        self.import_button.pack(pady=10)

        self.start_button = tk.Button(master, text="Start", command=self.start_conversion, state=tk.DISABLED)
        self.start_button.pack(pady=10)

        self.cancel_button = tk.Button(master, text="Cancel", command=self.cancel_conversion, state=tk.DISABLED)
        self.cancel_button.pack(pady=10)

        self.progress = ttk.Progressbar(master, length=200, mode='determinate')
        self.progress.pack(pady=10)

        self.time_label = tk.Label(master, text="Estimated time left: --:--")
        self.time_label.pack(pady=5)

        self.open_button = tk.Button(master, text="Open", command=self.open_file, state=tk.DISABLED)
        self.open_button.pack(pady=10)

    def import_pdf(self):
        self.pdf_file = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
        if self.pdf_file:
            self.start_button['state'] = tk.NORMAL

    def start_conversion(self):
        self.start_button['state'] = tk.DISABLED
        self.import_button['state'] = tk.DISABLED
        self.cancel_button['state'] = tk.NORMAL
        self.cancel_flag.clear()
        self.conversion_thread = threading.Thread(target=self.convert_pdf_to_excel, daemon=True)
        self.conversion_thread.start()

    def convert_pdf_to_excel(self):
        pdf_content = ""
        start_time = time.time()
        with open(self.pdf_file, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            total_pages = len(pdf_reader.pages)
            for i, page in enumerate(pdf_reader.pages):
                if self.cancel_flag.is_set():
                    break
                pdf_content += page.extract_text()
                self.update_progress((i + 1) / total_pages * 100, start_time, i + 1, total_pages)
                time.sleep(0.1)  # Simulate longer processing time

        if not self.cancel_flag.is_set():
            # Print the entire PDF content
            print("Full PDF content:")
            print(pdf_content)
            print("\n" + "="*50 + "\n")

            self.excel_file = os.path.splitext(self.pdf_file)[0] + ".xlsx"
            wb = Workbook()
            ws = wb.active
            
            # Set up column headers
            headers = ["Data", "Sistema/User", "Cliente", "Quadra", "Horário", "Valor", "Valor Recebido", 
                       "Valor Reserva", "Recebido Reserva", "Valor Materiais", "Recebido Materiais",
                       "Data Pagamento", "Forma Pgto", "Cliente Pagamento", "Valor Pagamento", "Usuário Pagamento", "Comentário"]
            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col, value=header)

            # Split the content into reservations
            reservations = re.split(r'\n(?=\d{2}/\d{2}/\d{4})', pdf_content)
            
            for row, reservation in enumerate(reservations, start=2):
                # Extract information using regex
                data = re.search(r'(\d{2}/\d{2}/\d{4})', reservation)
                sistema_user = re.search(r'(\d{2}/\d{2}/\d{4})\s+(SISTEMA|USER.*?)\s', reservation)
                cliente = re.search(r'(?:SISTEMA|USER.*?)\s+(.*?)\s+\d-', reservation)
                quadra = re.search(r'(\d-.*?)\s+\d{2}:', reservation)
                horario = re.search(r'(\d{2}:\d{2}\s*ATÉ\s*\d{2}:\d{2})', reservation)
                valor = re.search(r'R\$\s*(\d+,\d{2})', reservation)
                valor_recebido = re.search(r'R\$\s*(\d+,\d{2})\s*VALOR RESERVA:', reservation)
                valor_reserva = re.search(r'VALOR RESERVA:\s*R\$\s*(\d+,\d{2})', reservation)
                recebido_reserva = re.search(r'RECEBIDO RESERVA:\s*R\$\s*(\d+,\d{2})', reservation)
                valor_materiais = re.search(r'VALOR MATERIAIS:\s*R\$\s*(\d+,\d{2})', reservation)
                recebido_materiais = re.search(r'RECEBIDO MATERIAIS:\s*R\$\s*(\d+,\d{2})', reservation)
                data_pagamento = re.search(r'DATA PAGAMENTO\s+(\d{2}/\d{2}/\d{4}\s+[ÀA]S\s+\d{2}:\d{2}:\d{2})', reservation)
                forma_pgto = re.search(r'FORMA PGTO\.\s+(.*?)\s+CLIENTE', reservation)
                cliente_pagamento = re.search(r'CLIENTE\s+(.*?)\s+VALOR', reservation)
                valor_pagamento = re.search(r'VALOR\s+R\$\s*(\d+,\d{2})', reservation)
                usuario_pagamento = re.search(r'USUÁRIO\s+(.*?)\s+COMENTÁRIO:', reservation)
                comentario = re.search(r'COMENTÁRIO:\s*(.*?)$', reservation, re.MULTILINE | re.DOTALL)

                # Write extracted data to Excel
                ws.cell(row=row, column=1, value=data.group(1) if data else None)
                ws.cell(row=row, column=2, value=sistema_user.group(1) if sistema_user else None)
                ws.cell(row=row, column=3, value=cliente.group(1) if cliente else None)
                ws.cell(row=row, column=4, value=quadra.group(1) if quadra else None)
                ws.cell(row=row, column=5, value=horario.group(1) if horario else None)
                ws.cell(row=row, column=6, value=valor.group(1) if valor else None)
                ws.cell(row=row, column=7, value=valor_recebido.group(1) if valor_recebido else None)
                ws.cell(row=row, column=8, value=valor_reserva.group(1) if valor_reserva else None)
                ws.cell(row=row, column=9, value=recebido_reserva.group(1) if recebido_reserva else None)
                ws.cell(row=row, column=10, value=valor_materiais.group(1) if valor_materiais else None)
                ws.cell(row=row, column=11, value=recebido_materiais.group(1) if recebido_materiais else None)
                ws.cell(row=row, column=12, value=data_pagamento.group(1) if data_pagamento else None)
                ws.cell(row=row, column=13, value=forma_pgto.group(1) if forma_pgto else None)
                ws.cell(row=row, column=14, value=cliente_pagamento.group(1) if cliente_pagamento else None)
                ws.cell(row=row, column=15, value=valor_pagamento.group(1) if valor_pagamento else None)
                ws.cell(row=row, column=16, value=usuario_pagamento.group(1) if usuario_pagamento else None)
                ws.cell(row=row, column=17, value=comentario.group(1) if comentario else None)

            wb.save(self.excel_file)

        self.master.after(0, self.conversion_complete)

    def update_progress(self, value, start_time, current_page, total_pages):
        self.master.after(0, self.progress.config, {"value": value})
        
        elapsed_time = time.time() - start_time
        pages_left = total_pages - current_page
        estimated_time_left = (elapsed_time / current_page) * pages_left if current_page > 0 else 0
        self.master.after(0, self.time_label.config, {"text": f"Estimated time left: {estimated_time_left:.0f} seconds"})

    def cancel_conversion(self):
        self.cancel_flag.set()
        self.cancel_button['state'] = tk.DISABLED

    def conversion_complete(self):
        self.progress['value'] = 100
        self.start_button['state'] = tk.NORMAL
        self.import_button['state'] = tk.NORMAL
        self.cancel_button['state'] = tk.DISABLED
        self.time_label.config(text="Conversion complete")
        self.open_button['state'] = tk.NORMAL

    def open_file(self):
        if self.excel_file and os.path.exists(self.excel_file):
            if os.name == 'nt':  # For Windows
                os.startfile(self.excel_file)
            else:  # For macOS and Linux
                subprocess.call(('xdg-open', self.excel_file))

if __name__ == "__main__":
    root = tk.Tk()
    app = PDFConverter(root)
    root.mainloop()
